#!/usr/bin/env python3
"""Export i18n translations from src/i18n.js to Excel for translator review."""

import json, re, os

# Parse the i18n.js file
with open("src/i18n.js", "r") as f:
    content = f.read()

# Extract en and es blocks using regex
def extract_dict(content, lang):
    # Find the block for this language
    pattern = rf"  {lang}: \{{(.*?)\n  \}}"
    match = re.search(pattern, content, re.DOTALL)
    if not match:
        return {}
    block = match.group(1)
    result = {}
    for line in block.split("\n"):
        line = line.strip()
        if not line or line.startswith("//"):
            continue
        # Match key: 'value' or key: "value"
        m = re.match(r"(\w+):\s*['\"](.+?)['\"],?\s*$", line)
        if m:
            result[m.group(1)] = m.group(2)
        # Handle multi-word keys or escaped quotes
        m2 = re.match(r"(\w+):\s*[\"'](.+)[\"'],?\s*$", line)
        if not m and m2:
            result[m2.group(1)] = m2.group(2)
    return result

en = extract_dict(content, "en")
es = extract_dict(content, "es")

# Get all unique keys
all_keys = sorted(set(list(en.keys()) + list(es.keys())))

# Categorize keys by prefix/pattern
def categorize(key):
    if key.startswith("nav_"): return "Navigation"
    if key.startswith("chip_"): return "Filter Chips"
    if key.startswith("step"): return "Quote Flow"
    if key.startswith("ph_"): return "Placeholders"
    if key.startswith("cx_") or key.startswith("risk_"): return "Complexity & Risk"
    if key.startswith("map_"): return "Map"
    if key.startswith("bd_"): return "Formula Breakdown"
    if key.startswith("tip_"): return "Settings Tooltips"
    if key.startswith("err_"): return "Error Messages"
    if key.startswith("status_"): return "Status Labels"
    if key.startswith("plan_") or key in ("upgrade_to_pro","quotes_used"): return "Plans"
    if key.startswith("top_"): return "Business Tab"
    if key in ("email","password","confirm_password","log_in","create_account",
               "forgot_password","no_account","create_one","already_account",
               "log_in_link","back_to_site","send_reset","reset_password_title",
               "set_new_password","update_password","sign_in_subtitle",
               "create_account_subtitle","check_email_confirm"): return "Auth"
    if key in ("revenue","performance","quoted","quoted_sub","accepted_sub",
               "pending_sub","close_rate_pct","close_rate_sub","margin_sub",
               "day","week","month","year","today","this_week","this_month",
               "this_year","total_quotes","drafts","revenue_accepted",
               "total_quoted","accepted_revenue","pending_revenue",
               "close_rate","close_rate_detail","jobs_completed",
               "implied_hourly","implied_margin","business_title"): return "Business Tab"
    if key in ("company_logo","upload_logo","replace_logo","remove_logo",
               "company_name_label","company_phone","company_email",
               "save_settings","reset_defaults","log_out","auto_save_note",
               "formula_defaults","business_info","settings_title","language",
               "follow_up_reminders","follow_up_label","remind_after",
               "mow_rate_label","trim_rate_label","equipment_cost_label",
               "hourly_rate_label","minimum_bid_label","profit_margin_label",
               "quote_valid_days","complexity_default","risk_default",
               "logo_saved","logo_processing","reset_confirm"): return "Settings"
    if key in ("client_name","client_phone","client_email","property_address",
               "returning_client","client_info","job_address","client_label",
               "contact","clients_title","search_clients","no_clients",
               "edit_client","delete_client","search_name_phone",
               "no_clients_subtitle","existing_client_note",
               "last_measurements","quote_history","delete_client_title",
               "confirm_delete_all","confirm_final","are_you_sure",
               "delete_warn"): return "Clients"
    if key in ("service_type","one_time","recurring_service","frequency",
               "weekly","biweekly","monthly","is_recurring"): return "Recurring"
    if key in ("mark_visit_complete","visit_completed_on","next_visit",
               "keep_original","calculate_from_completion","schedule_manually",
               "end_season","mark_complete","set_next_visit","set_date",
               "season_complete","next_visit_tbd","visit","next_label",
               "service_label","overdue","due_today"): return "Visit Tracking"
    if key in ("attachments","add_photo","upload_file"): return "Attachments"
    if key in ("new_quote","all","draft","sent","accepted","declined",
               "recurring","follow_up","search_quotes","no_quotes",
               "start_first_quote","no_match","no_client"): return "Home Screen"
    if key in ("next","back","next_btn","back_btn","save","cancel",
               "delete","duplicate","copied","saving_label",
               "delete_confirm","new_quote_title","edit_quote_title",
               "new_revision","quote_details","step_of","of"): return "General UI"
    if key in ("send_quote","save_draft","resend_quote","download_pdf",
               "mark_accepted","mark_declined","edit_quote","cancel_service",
               "quote_summary","lawn_area","perimeter","use_measurements",
               "clear_all","undo_point","close_shape","area",
               "job_complexity","site_risk","discount","crew_and_time",
               "estimated_time","calculated_bid","tap_to_override",
               "breakdown","final_bid","notes","job_details",
               "formula_breakdown","quote_thread","name_required",
               "address_label","crew_label","est_time","created",
               "sent_date","expires","not_sent_label","none",
               "parallel","sequential","tap_override","reset_formula",
               "min_bid_applied","switch_to_manual","workers","worker",
               "expired_label","save_to_client","save_to_client_desc",
               "signup_pro_banner","trial_14"): return "Quote Flow"
    if key in ("mow_area","trim_perimeter","equipment","subtotal_costs",
               "complexity","risk","profit_margin"): return "Formula"
    if key in ("map_tab","manual_tab"): return "Measurements"
    return "Other"

# Try to use openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "LawnBid Translations"

# Header styling
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="15803d", end_color="15803d", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")

# Headers
headers = ["Key", "Category", "English", "Spanish", "Notes"]
widths = [30, 20, 50, 50, 30]
for col, (header, width) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap
    ws.column_dimensions[chr(64 + col)].width = width

# Data rows
for i, key in enumerate(all_keys, 2):
    ws.cell(row=i, column=1, value=key).alignment = wrap
    ws.cell(row=i, column=2, value=categorize(key)).alignment = wrap
    ws.cell(row=i, column=3, value=en.get(key, "")).alignment = wrap
    ws.cell(row=i, column=4, value=es.get(key, "⚠ MISSING")).alignment = wrap
    ws.cell(row=i, column=5, value="").alignment = wrap

# Freeze top row
ws.freeze_panes = "A2"

# Save
out_path = "dist/lawnbid-translations.xlsx"
os.makedirs("dist", exist_ok=True)
wb.save(out_path)
print(f"✓ Saved {out_path} with {len(all_keys)} translation keys")
